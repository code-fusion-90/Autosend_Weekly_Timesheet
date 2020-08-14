import openpyxl
from datetime import datetime
from datetime import timedelta
import shutil
import sendmail

#Updates G30 column with a day ending Sunday every week. Adds +7 to previous week sunday.
#E16-E20 define activities. Date formula for B16-B20 done in excel based on date in G30.
#Errors not handled. Work in progress
path = r"/Users/defaultuser/OneDrive"
filename = r"Time_sheet_Format.xlsx"

def main():
    theFile = openpyxl.load_workbook(path+filename)
    print(theFile.sheetnames)
    currentSheet = theFile['Time sheet format']
    date_time_str = currentSheet['G30'].value
    #date_time_obj = datetime. strptime(date_time_str, '%d/%m/%y %H:%M:%S')
    new_time = date_time_str + timedelta(days=7)
    new_time_str = new_time.strftime("%d %b %Y ")
    path_new_excel = path+"Time_sheet_Format_"+new_time_str+".xlsx"
    shutil.copy(path+filename, path_new_excel)
    newFile = openpyxl.load_workbook(path_new_excel)
    newSheet = newFile['Time sheet format']
    newSheet['G30'] = new_time
    newSheet['E16'] = 'Support' #mon
    newSheet['E17'] = 'Support' #tue
    newSheet['E18'] = 'Support' #wed
    newSheet['E19'] = 'Support' #thur
    newSheet['E20'] = 'Support' #fri
    newFile.save(path_new_excel)
    currentSheet['G30'] = new_time
    theFile.save(path+filename)
    sendmail.email(path_new_excel)
    #print(new_time)

main()
